import pyautogui
import openpyxl
import pyperclip

workbook = openpyxl.load_workbook(r'C:\Users\Lucas\OneDrive\Área de Trabalho\Projeto automação de cadastro de produtos\produtos.xlsx')
sheet_produtos = workbook['produtos']
max_linha = sheet_produtos.max_row
for linha in sheet_produtos.iter_rows(min_row=2, max_row=max_linha):
    produto = linha[0].value
    fornecedor = linha[1].value
    categoria = linha[2].value
    quantidade = linha[3].value
    valor_unitario = linha[4].value
    notificar_venda = linha[5].value
    pyautogui.click(1235,523, duration=1)
    pyautogui.write(produto)
    
    pyautogui.click(1549,522, duration=1)
    pyautogui.write(fornecedor)

    pyautogui.click(1179,640, duration=1)
    pyperclip.copy(categoria)
    pyautogui.hotkey('ctrl','v')

    pyautogui.click(1549,637, duration=1)
    pyperclip.copy(valor_unitario)
    pyautogui.hotkey('ctrl','v')  

    if notificar_venda == "Sim":
        pyautogui.click(1099,752, duration=1)
    elif notificar_venda == "Não":
        pyautogui.click(1233,752, duration=1)    

    pyautogui.click(1223,837, duration =1)   
    pyautogui.click(1656,207, duration=1)
